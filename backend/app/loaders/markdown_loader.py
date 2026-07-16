"""Markdown / TXT / HTML / CSV / XLSX 通用文本加载器。"""

from __future__ import annotations

from pathlib import Path

from app.loaders.base import BaseLoader, LoadedDocument, LoadedPage


class MarkdownLoader(BaseLoader):
     extensions = (".md", ".markdown")

     def load(self, path: str | Path) -> LoadedDocument:
         p = Path(path)
         text = p.read_text(encoding="utf-8", errors="ignore")
         pages = self._split_by_heading(text)
         return LoadedDocument(
             text=text,
             pages=pages,
             metadata={"char_count": len(text)},
             source=str(p),
             ext=p.suffix.lower(),
         )

     @staticmethod
     def _split_by_heading(text: str) -> list[LoadedPage]:
         pages: list[LoadedPage] = []
         current_title = None
         buf: list[str] = []
         for line in text.splitlines():
             stripped = line.strip()
             if stripped.startswith("#"):
                 if buf:
                     pages.append(
                         LoadedPage(text="\n".join(buf).strip(), section=current_title)
                     )
                     buf = []
                 current_title = stripped.lstrip("#").strip()
                 buf.append(stripped)
             else:
                 buf.append(line)
         if buf:
             pages.append(LoadedPage(text="\n".join(buf).strip(), section=current_title))
         return pages


class TextLoader(BaseLoader):
     extensions = (".txt", ".log")

     def load(self, path: str | Path) -> LoadedDocument:
         p = Path(path)
         text = p.read_text(encoding="utf-8", errors="ignore")
         return LoadedDocument(
             text=text,
             pages=[LoadedPage(text=text, page=1)] if text.strip() else [],
             metadata={"char_count": len(text)},
             source=str(p),
             ext=p.suffix.lower(),
         )


class HtmlLoader(BaseLoader):
     extensions = (".html", ".htm")

     def load(self, path: str | Path) -> LoadedDocument:
         from bs4 import BeautifulSoup

         p = Path(path)
         html = p.read_text(encoding="utf-8", errors="ignore")
         soup = BeautifulSoup(html, "html.parser")
         for tag in soup(["script", "style", "nav", "footer", "noscript"]):
             tag.decompose()
         title = (soup.title.string.strip() if soup.title and soup.title.string else None)
         text = soup.get_text(separator="\n", strip=True)
         return LoadedDocument(
             text=text,
             pages=[LoadedPage(text=text, page=1, section=title)],
             metadata={"title": title},
             source=str(p),
             ext=p.suffix.lower(),
         )


class CsvLoader(BaseLoader):
     extensions = (".csv",)

     def load(self, path: str | Path) -> LoadedDocument:
         import csv

         p = Path(path)
         with p.open(encoding="utf-8", errors="ignore", newline="") as f:
             reader = csv.reader(f)
             rows = [row for row in reader if any(c.strip() for c in row)]

         if not rows:
             return LoadedDocument(
                 text="",
                 pages=[],
                 metadata={"row_count": 0},
                 source=str(p),
                 ext=p.suffix.lower(),
             )

         header, *body = rows
         lines = [" | ".join(header)]
         for row in body:
             lines.append(" | ".join(row))
         text = "\n".join(lines)

         return LoadedDocument(
             text=text,
             pages=[LoadedPage(text=text, page=1, metadata={"header": header})],
             metadata={"row_count": len(body), "col_count": len(header)},
             source=str(p),
             ext=p.suffix.lower(),
         )


class XlsxLoader(BaseLoader):
     extensions = (".xlsx",)

     def load(self, path: str | Path) -> LoadedDocument:
         from openpyxl import load_workbook

         p = Path(path)
         wb = load_workbook(filename=str(p), read_only=True, data_only=True)
         pages: list[LoadedPage] = []
         for sheet_name in wb.sheetnames:
             ws = wb[sheet_name]
             lines: list[str] = [f"# Sheet: {sheet_name}"]
             for row in ws.iter_rows(values_only=True):
                 cells = [str(c).strip() if c is not None else "" for c in row]
                 if any(cells):
                     lines.append(" | ".join(cells))
             if len(lines) > 1:
                 pages.append(LoadedPage(text="\n".join(lines), section=sheet_name))
         wb.close()

         full = "\n\n".join(p.text for p in pages)
         return LoadedDocument(
             text=full,
             pages=pages,
             metadata={"sheet_count": len(pages)},
             source=str(p),
             ext=p.suffix.lower(),
         )
